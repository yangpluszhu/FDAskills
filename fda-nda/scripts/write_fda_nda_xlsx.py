#!/usr/bin/env python3
"""
Write curated FDA approval rows to an XLSX workbook.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


COLUMN_ORDER = [
    ("drug_type_zh", "药物类型"),
    ("generic_name", "通用名"),
    ("brand_name", "商品名"),
    ("manufacturer_zh", "生产商"),
    ("indication_zh", "适应症"),
    ("approval_date", "批准日期"),
    ("significance_zh", "上市意义或价值"),
    ("fda_source_zh", "FDA官方来源"),
]

COLUMN_WIDTHS = {
    "A": 12,
    "B": 24,
    "C": 24,
    "D": 24,
    "E": 56,
    "F": 14,
    "G": 52,
    "H": 54,
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Write fdaNDA.xlsx from curated JSON rows.")
    parser.add_argument("--input", required=True, help="Curated JSON input path.")
    parser.add_argument("--output", required=True, help="XLSX output path.")
    return parser.parse_args()


def load_rows(path: Path) -> list[dict[str, Any]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict) and isinstance(payload.get("records"), list):
        return payload["records"]
    raise ValueError("Input JSON must be a list or an object with a 'records' list.")


def coerce_source(row: dict[str, Any]) -> str:
    if row.get("fda_source_zh"):
        return str(row["fda_source_zh"])
    urls = row.get("source_urls") or []
    if isinstance(urls, list):
        return "\n".join(str(item) for item in urls)
    return ""


def build_workbook(rows: list[dict[str, Any]]) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "fdaNDA"

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for column_index, (_, header) in enumerate(COLUMN_ORDER, start=1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, row in enumerate(rows, start=2):
        row_data = {
            "drug_type_zh": row.get("drug_type_zh", ""),
            "generic_name": row.get("generic_name", ""),
            "brand_name": row.get("brand_name", ""),
            "manufacturer_zh": row.get("manufacturer_zh", ""),
            "indication_zh": row.get("indication_zh", ""),
            "approval_date": row.get("approval_date", ""),
            "significance_zh": row.get("significance_zh", ""),
            "fda_source_zh": coerce_source(row),
        }
        for column_index, (field_name, _) in enumerate(COLUMN_ORDER, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=row_data[field_name])
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if field_name == "fda_source_zh" and isinstance(row.get("source_urls"), list):
                urls = row["source_urls"]
                if len(urls) == 1:
                    cell.hyperlink = urls[0]
                    cell.style = "Hyperlink"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for column_letter, width in COLUMN_WIDTHS.items():
        sheet.column_dimensions[column_letter].width = width

    for row_cells in sheet.iter_rows(min_row=2):
        max_lines = max(len(str(cell.value).splitlines()) if cell.value else 1 for cell in row_cells)
        sheet.row_dimensions[row_cells[0].row].height = max(20, max_lines * 16)

    return workbook


def main() -> int:
    args = parse_args()
    input_path = Path(args.input).resolve()
    output_path = Path(args.output).resolve()

    rows = load_rows(input_path)
    workbook = build_workbook(rows)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(f"Wrote {len(rows)} rows to {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
